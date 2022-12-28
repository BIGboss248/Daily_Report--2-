import re
import os
import datetime
from datetime import date
import string
import tkinter
from tkinter import filedialog
import pandas as pd
import openpyxl as xl
from openpyxl import formatting, styles, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.table import Table, TableStyleInfo


def find_commodity_price_row(Platts_String: str, commodity_symbol: str, second_pattern='\s+\d+.+') -> re.match:
    """ Finds the row corosponding to the commodity inside the daily report and returns a re.match"""
    # the pattern is based on the symbol so in the program what the  function will recive is the symbol
    # corosponding to the needed commodity
    pattern = "".join([rf'{commodity_symbol}', second_pattern])
    compiled_pattern = re.compile(pattern)
    matches = re.search(pattern=compiled_pattern, string=Platts_String)
    return matches


def extract_numbers(out_match: re.match, index: int) -> dict:
    """gets the file  and commodity symbol and uses the find_commodity_price_row gets the row of the commodity inside the daily report and
    removes the name symbol and spaces of the row and retruns a dictionary contaning the info"""
    # creating a list of english letters to remove the name and symbol
    alphabet = list(string.ascii_letters)
    # extracting the string from re,match and turning it into a list of strings
    in_list = out_match.group().split(" ")
    # removing the space betwean the numbers and name and symbol
    in_list = list(dict.fromkeys(in_list))
    # creating a list of items that should be removed from the original re.match string
    remove_list = []
    for i in in_list:
        if i.isalpha() or i == "" or '%' in i:
            remove_list.append(i)
    for i in in_list:
        for a in alphabet:
            if a in i:
                remove_list.append(i)
                break
    # removing duplicates from remove_list
    remove_list = list(dict.fromkeys(remove_list))
    # removing the items in in_list that are in remove_list
    for i in remove_list:
        in_list.remove(i)
    # in special cases there will be a '.' that will cause error so we check for it and  delete it if its there
    if '.' in in_list:
        in_list.remove('.')
    # turning the list of strings into float
    in_list = [float(x) for x in in_list]
    if len(in_list) == 0:
        in_list.append('-')
        in_list.append('-')
        in_list.append('-')

    if index == 2:
        result = {'Price': in_list[0]}
    if index == 3:
        result = {'Price': in_list[0], 'Change': in_list[1]}
    if index == 4:
        result = {'Price': in_list[0],
                  'Change': in_list[1], 'Change %': in_list[2]}
    return result


def final_report(Platts_String: str, commodity_dict: dict, needed_numbers: int,
                 second_pattern='\s+\d+.+') -> pd.DataFrame:
    """ Gets the inputs for find_commodity_price_row and extract_numbers then using these functions returns a pandas dataframe
        of the commodity price and attributes"""
    # dataframe_index is used to create dataframes from the dictionaries and has no purpose other than that
    dataframe_index = 0
    # The final dataframe created form the commodity_dict to publish
    complete_dataframe = pd.DataFrame()
    for commodity_name in list(commodity_dict.keys()):
        # commodity_symbol is used to identify the relevent attributes for the commodity_name
        commodity_symbol = commodity_dict[commodity_name]['symbol']
        # commodity_attributes_dict is the constructor dictionary for commodity_df
        commodity_attributes_dict = commodity_dict[commodity_name]['attributes']
        match = find_commodity_price_row(
            Platts_String, commodity_symbol, second_pattern)
        numbers = extract_numbers(match, needed_numbers)
        # assigning commodity name which is used as needed_numbers
        commodity_attributes_dict['Commodity'] = commodity_name
        # assigning extracted numbers
        commodity_attributes_dict['Price'] = numbers['Price']
        if needed_numbers == 3:
            commodity_attributes_dict['Change'] = numbers['Change']
        if needed_numbers == 4:
            commodity_attributes_dict['Change'] = numbers['Change']
            commodity_attributes_dict['Change %'] = numbers['Change %']
        # constructing df of the commodity_name to later add it to the complete_dataframe
        commodity_df = pd.DataFrame(
            commodity_attributes_dict, index=[dataframe_index])
        commodity_df.set_index('Commodity', inplace=True)
        complete_dataframe = pd.concat([complete_dataframe, commodity_df])
        # adding 1 to dataframe index to ditinguish diffrent commodity_name rows
        dataframe_index = dataframe_index + 1
    return complete_dataframe


def translate_report(dataframe: pd.DataFrame, persian_dict: dict):
    new_column_names = []
    new_indexes = {}
    for column_name in dataframe.columns:
        persian_word = persian_dict[column_name]
        new_column_names.append(persian_word)
    dataframe.columns = new_column_names
    new_index = persian_dict[dataframe.index.name]
    dataframe.index.name = new_index
    for i in range(len(dataframe.index)):
        translated = persian_dict[dataframe.index[i]]
        new_indexes[dataframe.index[i]] = translated
    dataframe.rename(index=new_indexes, inplace=True)
    return dataframe


def get_Volume_Issue_Date(Platts_String: str) -> dict:
    """gets the file string returns a dictionary first one will be the volume second the issue number and third a date object"""
    pattern = re.compile(r'Volume.+')
    match = re.search(pattern, Platts_String)
    match_string = match.group()
    Volume = int(match_string.split('/')[0].split(' ')[1])
    Issue = int(match_string.split('/')[1].split(' ')[2])
    month = match_string.split('/')[2].split(' ')[1]
    month_dict = {'January': 1, 'February': 2, 'March': 3,
                  'April': 4, 'May': 5, 'June': 6, 'July': 7,
                  'August': 8, 'September': 9, 'October': 10,
                  'November': 11, 'December': 12}
    day = int(match_string.split('/')[2].split(' ')[2].replace(',', ""))
    year = int(match_string.split('/')[2].split(' ')[3])
    date_stamp = datetime.date(year, month_dict[month], day)
    return {'Volume': Volume, 'Issue': Issue, 'date_stamp': date_stamp}


def export_to_excel(excel_file_address: str, dataframe_dict: dict):
    """ With a openpyxl.writer writes the dataframes to excel worksheets"""
    excel_writer = pd.ExcelWriter(excel_file_address, engine='openpyxl')
    for dataframe in dataframe_dict:
        dataframe_dict[dataframe].to_excel(excel_writer, sheet_name=dataframe)
    excel_writer.close()


def excel_set_font(excel_file_address: str, font=Font(name='IRNazanin', size=16)):
    """Takes the address of an excel file Adjusts font"""
    wb = xl.load_workbook(excel_file_address)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for i in range(1, ws.max_row+1):
            for j in range(1, ws.max_column+1):
                selected_cell = ws.cell(row=i, column=j)
                selected_cell.font = font
    wb.save(excel_file_address)


def excel_set_alignment(excel_file_address: str, alignment=Alignment(horizontal='center', vertical='center')):
    """Takes the address of an excel file Adjusts alignment"""
    wb = xl.load_workbook(excel_file_address)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for i in range(1, ws.max_row+1):
            for j in range(1, ws.max_column+1):
                selected_cell = ws.cell(row=i, column=j)
                selected_cell.alignment = alignment
    wb.save(excel_file_address)


def excel_set_border(excel_file_address: str,
                     border=Border(left=Side(border_style="thin", color='000000'),
                                   right=Side(border_style="thin",
                                              color='000000'),
                                   top=Side(border_style="thin",
                                            color='000000'),
                                   bottom=Side(border_style="thin", color='000000'))):
    """Takes the address of an excel file Adjusts border"""
    wb = xl.load_workbook(excel_file_address)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for i in range(1, ws.max_row+1):
            for j in range(1, ws.max_column+1):
                selected_cell = ws.cell(row=i, column=j)
                selected_cell.border = border
    wb.save(excel_file_address)


def excel_set_tables(excel_file_address: str, Table_Style=TableStyleInfo(name="TableStyleMedium12")):
    """ Takes the address of an excel file and defines Tables with styles"""
    wb = xl.load_workbook(excel_file_address)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Create Table and set a table style
        tab = Table(displayName=sheet,
                    ref=rf'A1:{string.ascii_uppercase[ws.max_column-1]}{ws.max_row}', tableStyleInfo=Table_Style)
        ws.add_table(tab)
    wb.save(excel_file_address)


def excel_set_number_formats(excel_file_address: str, percentage_list=['Fe', 'silica', 'moisture', 'alumina', 'phosphorus', 'sulfur'],
                             currency_list=['Price', 'Change', 'Change %'], currency_format='"$"#,##0.00_-', percentage_format='0.00%'):
    """ Gets excel file address and two lists containing  the header names of the values that should be formatted as a percentage
        or currency the default format for currency is dolor it searches each sheet for the header and if the item in the list is 
        not in the sheet function simply ignores it and moves to the next sheet"""
    wb = xl.load_workbook(excel_file_address)
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        currency_column_num = {}
        for j in range(1, ws.max_column+1):
            selected_cell = ws.cell(row=1, column=j)
            if selected_cell.value in currency_list:
                currency_column_num[selected_cell.value] = j
        for column in list(currency_column_num.keys()):
            for i in range(2, ws.max_row+1):
                selected_cell = ws.cell(
                    row=i, column=currency_column_num[column])
                selected_cell.number_format = currency_format

        percentage_column_num = {}
        for j in range(1, ws.max_column+1):
            selected_cell = ws.cell(row=1, column=j)
            if selected_cell.value in percentage_list:
                percentage_column_num[selected_cell.value] = j
        for column in list(percentage_column_num.keys()):
            for i in range(2, ws.max_row+1):
                selected_cell = ws.cell(
                    row=i, column=percentage_column_num[column])
                if selected_cell.value != None:
                    selected_cell.value = selected_cell.value/100.00
                    selected_cell.number_format = percentage_format

    wb.save(excel_file_address)


def excel_set_column_width(excel_file_address: str):
    wb = xl.load_workbook(excel_file_address)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Set column width to auto size
        for column_letter in string.ascii_uppercase:
            max_width = 0
            for row in range(1, ws.max_row+1):
                row_len = len(str(ws[f'{column_letter}{row}'].value))
                if row_len > max_width:
                    max_width = row_len
            ws.column_dimensions[column_letter].width = max_width + 13
    wb.save(excel_file_address)


def excel_set_conditional_formatting(excel_file_address: str,rule_columns=['Change', 'Change %']):
    red_color = 'ffc7ce'
    red_color_font = '9c0103'
    green_color = 'C6EFCE'
    green_color_font = '006100'
    yellow_color = 'FFEB9C'
    yellow_color_font = '9C5700'
    red_font = styles.Font(color=red_color_font)
    red_fill = styles.PatternFill(
        start_color=red_color, end_color=red_color, fill_type='solid')
    green_font = styles.Font(color=green_color_font)
    green_fill = styles.PatternFill(
        start_color=green_color, end_color=green_color, fill_type='solid')
    yellow_font = styles.Font(color=yellow_color_font)
    yellow_fill = styles.PatternFill(
        start_color=yellow_color, end_color=yellow_color, fill_type='solid')
    wb = xl.load_workbook(excel_file_address)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rules_column_nums = []
        for j in range(1, ws.max_column+1):
            selected_cell = ws.cell(row=1, column=j)
            if selected_cell.value in rule_columns:
                rules_column_nums.append(j)
        for i in rules_column_nums:
            apply_range = rf"{string.ascii_uppercase[i-1]}2:{string.ascii_uppercase[i-1]}{ws.max_row}"
            # start from row 2 and column 2 to ignore headers and indexes
            ws.conditional_formatting.add(apply_range, formatting.rule.CellIsRule(
                operator='lessThan', formula=['0'], fill=red_fill, font=red_font))
            ws.conditional_formatting.add(apply_range, formatting.rule.CellIsRule(
                operator='lessThan', formula=['0'], fill=red_fill))
            ws.conditional_formatting.add(apply_range, formatting.rule.CellIsRule(
                operator='greaterThan', formula=['0'], fill=green_fill, font=green_font))
            ws.conditional_formatting.add(apply_range, formatting.rule.CellIsRule(
                operator='greaterThan', formula=['0'], fill=green_fill))
            ws.conditional_formatting.add(apply_range, formatting.rule.CellIsRule(
                operator='equal', formula=['0'], fill=yellow_fill, font=yellow_font))
            ws.conditional_formatting.add(apply_range, formatting.rule.CellIsRule(
                operator='equal', formula=['0'], fill=yellow_fill))
    wb.save(excel_file_address)


def excel_format(excel_file_address: str, font=Font(name='IRNazanin', size=16), alignment=Alignment(horizontal='center', vertical='center'),
                 border=Border(left=Side(border_style="thin", color='000000'), right=Side(border_style="thin", color='000000'),
                               top=Side(border_style="thin", color='000000'), bottom=Side(border_style="thin", color='000000')),
                 Table_Style=TableStyleInfo(name="TableStyleMedium12"), percentage_list=['Fe', 'silica', 'moisture', 'alumina', 'phosphorus', 'sulfur'],
                 currency_list=['Price', 'Change', 'Change %'], currency_format='"$"#,##0.00_-', percentage_format='0.00%',
                 rule_columns=['Change', 'Change %']):
    """ Takes the address of an excel file and Adjusts column width font and format alignment border and table style"""
    excel_set_font(excel_file_address, font)
    excel_set_alignment(excel_file_address, alignment)
    excel_set_border(excel_file_address, border)
    excel_set_tables(excel_file_address, Table_Style)
    excel_set_number_formats(excel_file_address, percentage_list=percentage_list, currency_list=currency_list,
                             currency_format=currency_format, percentage_format=percentage_format)
    excel_set_column_width(excel_file_address)
    excel_set_conditional_formatting(excel_file_address,rule_columns)


# declare addresses
Platts_file_full_address = r'G:\Shared drives\Unlimited Drive\Scripts\Daily_Report (2)\Resources\Platts-text.txt'

# Open Platts file
Platts_Daily_Report_File = open(
    Platts_file_full_address, 'rt', encoding='utf-8')
Platts_Daily_Report_String = Platts_Daily_Report_File.read()
Platts_Daily_Report_File.close()

# list of commoditys
indexes = {
    'IODEX 62% Fe CFR North China': {'symbol': 'IODBZ00', 'attributes': {'Fe': 62, 'moisture': 8,
                                                                         'silica': 4, 'alumina': 2.25,
                                                                         'phosphorus': 0.02, 'sulfur': 0.02}},
    '65% Fe CFR North China': {'symbol': 'IOPRM00', 'attributes': {'Fe': 65, 'moisture': 8.5,
                                                                   'silica': 3.5, 'alumina': 1,
                                                                   'phosphorus': 0.075, 'sulfur': None}},
    '58% Fe CFR North China': {'symbol': 'IODFE00', 'attributes': {'Fe': 58, 'moisture': 10,
                                                                   'silica': 5, 'alumina': 4,
                                                                   'phosphorus': 0.075, 'sulfur': None}}}

lump = {
    'Lump outright': {'symbol': 'IOCLS00', 'attributes': {'Fe': 62, 'moisture': 4,
                                                          'silica': 3.5, 'alumina': 1.5,
                                                          'phosphorus': 0.075, 'sulfur': 0.02}}}

pellet = {
    'Weekly CFR China 65% Fe': {'symbol': 'IOBFC04', 'attributes': {'Fe': 65, 'alumina': 0.35,
                                                                    'silica': 5, 'phosphorus': 0.02,
                                                                    'sulfur': 0.003, 'CCS': 250}},
    'Daily CFR China 63% Fe spot fixed price assessment': {'symbol': 'IOCQR04', 'attributes': {'Fe': 64, 'alumina': 2.7,
                                                                                               'silica': 3.5, 'phosphorus': 0.08,
                                                                                               'sulfur': 0.008, 'CCS': 230}},
    'Atlantic Basin 65% Fe Blast Furnace pellet FOB Brazil': {'symbol': 'SB01095', 'attributes': {'Fe': 65, 'alumina': 0.5,
                                                                                                  'silica': 3, 'phosphorus': None,
                                                                                                  'sulfur': None, 'CCS': 275}},
    'Direct Reduction 67.5% Fe pellet premium (65% Fe basis)': {'symbol': 'IODBP00', 'attributes': {'Fe': 67.5, 'alumina': None,
                                                                                                    'silica': 1.5, 'phosphorus': None,
                                                                                                    'sulfur': None, 'CCS': 300}}}

ore_brands = {
    'Pilbara Blend Fines (PBF) CFR Qingdao': {'symbol': 'IOPBQ00', 'attributes': {'Fe': 62, 'moisture': 8,
                                                                                  'silica': 4, 'alumina': 2.25,
                                                                                  'phosphorus': 0.02, 'sulfur': 0.02}},
    'Brazilian Blend Fines (BRBF) CFR Qingdao': {'symbol': 'IOBBA00', 'attributes': {'Fe': 62, 'moisture': 8,
                                                                                     'silica': 4, 'alumina': 2.25,
                                                                                     'phosphorus': 0.02, 'sulfur': 0.02}},
    'Newman High Grade Fines (NHGF) CFR Qingdao': {'symbol': 'IONHA00', 'attributes': {'Fe': 62, 'moisture': 8,
                                                                                       'silica': 4, 'alumina': 2.25,
                                                                                       'phosphorus': 0.02, 'sulfur': 0.02}},
    'Mining Area C Fines (MACF) CFR Qingdao': {'symbol': 'IOMAA00', 'attributes': {'Fe': 62, 'moisture': 8,
                                                                                   'silica': 4, 'alumina': 2.25,
                                                                                   'phosphorus': 0.02, 'sulfur': 0.02}},
    'Jimblebar Fines (JMBF) CFR Qingdao': {'symbol': 'IOJBA00', 'attributes': {'Fe': 62, 'moisture': 8,
                                                                               'silica': 4, 'alumina': 2.25,
                                                                               'phosphorus': 0.02, 'sulfur': 0.02}},
    '57% Fe Yandi Fines (YDF) CFR Qingdao': {'symbol': 'IOYFA00', 'attributes': {'Fe': 57, 'moisture': 10,
                                                                                 'silica': 5, 'alumina': 4,
                                                                                           'phosphorus': 0.075, 'sulfur': None}}}

Asia_Pacific_coking_coal = {
    'HCC Peak Downs Region FOB Australia': {'symbol': 'HCCGA00', 'attributes': {}},
    'HCC Peak Downs Region CFR China': {'symbol': 'HCCGC00', 'attributes': {}},
    'HCC Peak Downs Region CFR India': {'symbol': 'HCCGI00', 'attributes': {}},
    'Premium Low Vol FOB Australia': {'symbol': 'PLVHA00', 'attributes': {}},
    'Premium Low Vol CFR China': {'symbol': 'PLVHC00', 'attributes': {}},
    'Premium Low Vol CFR India': {'symbol': 'PLVHI00', 'attributes': {}},
    'Low Vol HCC FOB Australia': {'symbol': 'HCCAU00', 'attributes': {}},
    'Low Vol HCC CFR China': {'symbol': 'HCCCH00', 'attributes': {}},
    'Low Vol HCC CFR India': {'symbol': 'HCCIN00', 'attributes': {}},
    'Low Vol PCI FOB Australia': {'symbol': 'MCLVA00', 'attributes': {}},
    'Low Vol PCI CFR China': {'symbol': 'MCLVC00', 'attributes': {}},
    'Low Vol PCI CFR India': {'symbol': 'MCLVI00', 'attributes': {}},
    'Mid Vol PCI FOB Australia': {'symbol': 'MCLAA00', 'attributes': {}},
    'Mid Vol PCI CFR China': {'symbol': 'MCLAC00', 'attributes': {}},
    'Mid Vol PCI CFR India': {'symbol': 'MCVAI00', 'attributes': {}},
    'Semi Soft FOB Australia': {'symbol': 'MCSSA00', 'attributes': {}},
    'Semi Soft CFR China': {'symbol': 'MCSSC00', 'attributes': {}},
    'Semi Soft CFR India': {'symbol': 'MCSSI00', 'attributes': {}}}

Asia_Pacific_brand_relativities_Premium_Low_Vol = {
    'Peak Downs FOB Australia': {'symbol': 'HCPDA00', 'attributes': {}},
    'Peak Downs CFR China': {'symbol': 'MCBAA00', 'attributes': {}},
    'Saraji FOB Australia': {'symbol': 'HCSAA00', 'attributes': {}},
    'Saraji CFR China': {'symbol': 'MCBAB00', 'attributes': {}},
    'Oaky North FOB Australia': {'symbol': 'HCOKA00', 'attributes': {}},
    'Oaky North CFR China': {'symbol': 'MCBAR00', 'attributes': {}},
    'Illawarra FOB Australia': {'symbol': 'HCIWA00', 'attributes': {}},
    'Illawarra CFR China': {'symbol': 'MCBAH00', 'attributes': {}},
    'Moranbah North FOB Australia': {'symbol': 'HCMOA00', 'attributes': {}},
    'Moranbah North CFR China': {'symbol': 'MCBAG00', 'attributes': {}},
    'Goonyella FOB Australia': {'symbol': 'HCGOA00', 'attributes': {}},
    'Goonyella CFR China': {'symbol': 'MCBAE00', 'attributes': {}},
    'Peak Downs North FOB Australia': {'symbol': 'HCPNA00', 'attributes': {}},
    'Peak Downs North CFR China': {'symbol': 'MCBAJ00', 'attributes': {}},
    'Goonyella C FOB Australia': {'symbol': 'HCGNA00', 'attributes': {}},
    'Goonyella C CFR China': {'symbol': 'MCBAI00', 'attributes': {}},
    'Riverside FOB Australia': {'symbol': 'HCRVA00', 'attributes': {}},
    'Riverside CFR China': {'symbol': 'MCRVR00', 'attributes': {}},
    'GLV FOB Australia': {'symbol': 'HCHCA00', 'attributes': {}},
    'GLV CFR China': {'symbol': 'MCBAF00', 'attributes': {}}}

Asia_Pacific_brand_relativities_Low_Vol_HCC = {
    'Lake Vermont HCC': {'symbol': 'MCBAN00', 'attributes': {}},
    'Carborough Downs': {'symbol': 'MCBAO00', 'attributes': {}},
    'Middlemount Coking': {'symbol': 'MCBAP00', 'attributes': {}},
    'Poitrel Semi Hard': {'symbol': 'MCBAQ00', 'attributes': {}}}

Dry_bulk_freight_assessments = {
    'Australia-China-Capesize': {'symbol': 'CDANC00', 'attributes': {}},
    'Australia-Rotterdam-Capesize': {'symbol': 'CDARN00', 'attributes': {}},
    'Australia-China-Panamax': {'symbol': 'CDBFA00', 'attributes': {}},
    'Australia-India-Panamax': {'symbol': 'CDBFAI0', 'attributes': {}},
    'USEC-India-Panamax': {'symbol': 'CDBUI00', 'attributes': {}},
    'USEC-Rotterdam-Panamax': {'symbol': 'CDBUR00', 'attributes': {}},
    'USEC-Brazil-Panamax': {'symbol': 'CDBUB00', 'attributes': {}},
    'US Mobile-Rotterdam-Panamax': {'symbol': 'CDMAR00', 'attributes': {}}
}


df_indexes = final_report(Platts_Daily_Report_String, indexes, 4)

df_lump = final_report(Platts_Daily_Report_String, lump, 3)

df_pellet = final_report(Platts_Daily_Report_String, pellet, 3)
# Adding the IODEX Price to Premiums
df_pellet.loc['Weekly CFR China 65% Fe', 'Price'] = df_pellet.loc['Weekly CFR China 65% Fe', 'Price'] + \
    df_indexes.loc['IODEX 62% Fe CFR North China', 'Price']
df_pellet.loc['Direct Reduction 67.5% Fe pellet premium (65% Fe basis)', 'Price'] = df_pellet.loc[
    'Direct Reduction 67.5% Fe pellet premium (65% Fe basis)', 'Price'] + df_indexes.loc['IODEX 62% Fe CFR North China', 'Price']

df_ore_brands = final_report(Platts_Daily_Report_String, ore_brands, 3)

df_Asia_Pacific_coking_coal = final_report(
    Platts_Daily_Report_String, Asia_Pacific_coking_coal, 3)

df_Asia_Pacific_brand_relativities_Premium_Low_Vol = final_report(
    Platts_Daily_Report_String, Asia_Pacific_brand_relativities_Premium_Low_Vol, 2)

df_Asia_Pacific_brand_relativities_Low_Vol_HCC = final_report(
    Platts_Daily_Report_String, Asia_Pacific_brand_relativities_Low_Vol_HCC, 2)

df_Dry_bulk_freight_assessments = final_report(
    Platts_Daily_Report_String, Dry_bulk_freight_assessments, 3, second_pattern='.+')

# List of data frames for exporting to excel file
dataframe_dict = {
    'indexes': df_indexes, 'lump': df_lump, 'pellet': df_pellet, 'ore_brands': df_ore_brands,
    'coking_coal': df_Asia_Pacific_coking_coal,
    'Premium_Coal': df_Asia_Pacific_brand_relativities_Premium_Low_Vol,
    'HCC_Coal': df_Asia_Pacific_brand_relativities_Low_Vol_HCC,
    'freight_assessments': df_Dry_bulk_freight_assessments}

translate_dict = {
    'Price': 'قیمت', 'Change': 'تغییر', 'Change %': 'درصد تغییر', 'Commodity': 'کالا',
    'Fe': 'آهن', 'moisture': 'رطوبت', 'silica': 'سیلیکا', 'alumina': 'آلومینا', 'phosphorus': 'فسفر', 'sulfur': 'سولفور', 'CCS': 'شاخص سختی',
    'IODEX 62% Fe CFR North China': 'ریزدانه تحویل به چین', '65% Fe CFR North China': 'ریزدانه تحویل به چین',
    '58% Fe CFR North China': 'ریزدانه تحویل به چین', 'Lump outright': 'درشت دانه', 'Weekly CFR China 65% Fe': 'گندله کوره تحویل به چین',
    'Daily CFR China 63% Fe spot fixed price assessment': 'گندله کوره تحویل به چین',
    'Atlantic Basin 65% Fe Blast Furnace pellet FOB Brazil': 'گندله کوره تحویل بندر برزیل',
    'Direct Reduction 67.5% Fe pellet premium (65% Fe basis)': 'گندله اسفنجی', 'Pilbara Blend Fines (PBF) CFR Qingdao': 'ریزدانه پیلبارا',
    'Brazilian Blend Fines (BRBF) CFR Qingdao': 'بلند(ترکیب) ریزدانه برزیل',
    'Newman High Grade Fines (NHGF) CFR Qingdao': 'گندله با کیفیت نیومن',
    'Mining Area C Fines (MACF) CFR Qingdao': 'ریزدانه مک', 'Jimblebar Fines (JMBF) CFR Qingdao': 'ریزدانه جیمبلبار',
    '57% Fe Yandi Fines (YDF) CFR Qingdao': 'ریزدانه یاندی', 'HCC Peak Downs Region FOB Australia': 'زغال سنگ متالورژی پیک داونز تحویل استرالیا',
    'HCC Peak Downs Region CFR China': 'زغال سنگ متالورژی پیک داونز تحویل چین', 'HCC Peak Downs Region CFR India': 'زغال سنگ متالورژی پیک داونز تحویل هند',
    'Premium Low Vol FOB Australia': 'زغال متالورژی با کیفیت با مواد فرار کم تحویل به استرالیا', 'Premium Low Vol CFR China': 'زغال متالورژی با کیفیت با مواد فرار کم تحویل به چین',
    'Premium Low Vol CFR India': 'زغال متالورژی با کیفیت با مواد فرار کم تحویل به هند', 'Low Vol HCC FOB Australia': 'زغال متالورژی استاندارد با مواد فرار کم تحویل به استرالیا',
    'Low Vol HCC CFR China': 'زغال متالورژی استاندارد با مواد فرار کم تحویل به چین', 'Low Vol HCC CFR India': 'زغال متالورژی استاندارد با مواد فرار کم تحویل به هند',
    'Low Vol PCI FOB Australia': 'زغال سنگ خام مواد فرار کم تحویل به استرالیا', 'Low Vol PCI CFR China': 'زغال سنگ خام مواد فرار کم تحویل به چین',
    'Low Vol PCI CFR India': 'زغال سنگ خام مواد فرار کم تحویل به هند', 'Mid Vol PCI FOB Australia': 'زغال سنگ خام مواد فرار متوسط تحویل به استرالیا',
    'Mid Vol PCI CFR China': 'زغال سنگ خام مواد فرار متوسط تحویل به چین', 'Mid Vol PCI CFR India': 'زغال سنگ خام مواد فرار متوسط تحویل به هند',
    'Semi Soft FOB Australia': 'زغال سنگ کک نیمه نرم تحویل به استرالیا', 'Semi Soft CFR China': 'زغال سنگ کک نیمه نرم تحویل به چین',
    'Semi Soft CFR India': 'زغال سنگ کک نیمه نرم تحویل به هند', 'Peak Downs FOB Australia': 'زغال سنگ مواد فرار کم پیک داونز تحویل به استرالیا',
    'Peak Downs CFR China': 'زغال سنگ مواد فرار کم پیک داونز تحویل به چین', 'Saraji FOB Australia': 'زغال سنگ مواد فرار کم سراجی تحویل به استرالیا',
    'Saraji CFR China': 'زغال سنگ مواد فرار کم سراجی تحویل به چین', 'Oaky North FOB Australia': 'زغال سنگ مواد فرار کم او کی نرث تحویل به استرالیا',
    'Oaky North CFR China': 'زغال سنگ مواد فرار کم او کی نرث تحویل به چین', 'Illawarra FOB Australia': 'زغال سنگ مواد فرار کم ایلاوارا تحویل به استرالیا',
    'Illawarra CFR China': 'زغال سنگ مواد فرار کم ایلاوارا تحویل به استرالیا',
    'Moranbah North FOB Australia': 'زغال سنگ مواد فرار کم مورانباه نرث تحویل به استرالیا',
    'Moranbah North CFR China': 'زغال سنگ مواد فرار کم مورانباه نرث تحویل به چین',
    'Goonyella FOB Australia': 'زغال سنگ مواد فرار کم گونیلا تحویل به استرالیا',
    'Goonyella CFR China': 'زغال سنگ مواد فرار کم گونیلا تحویل به چین',
    'Peak Downs North FOB Australia': 'زغال سنگ مواد فرار کم پیک داونز نرث تحویل به استرالیا',
    'Peak Downs North CFR China': 'زغال سنگ مواد فرار کم پیک داونز تحویل به چین',
    'Goonyella C FOB Australia': 'زغال سنگ مواد فرار کم گونیلا سی تحویل به استرالیا',
    'Goonyella C CFR China': 'زغال سنگ مواد فرار کم گونیلا سی تحویل به چین',
    'Riverside FOB Australia': 'زغال سنگ مواد فرار کم ریورساید تحویل به استرالیا',
    'Riverside CFR China': 'زغال سنگ مواد فرار کم ریورساید تحویل به چین',
    'GLV FOB Australia': 'زغال سنگ مواد فرار کم جی ال وی تحویل به استرالیا',
    'GLV CFR China': 'زغال سنگ مواد فرار کم جی ال وی تحویل به چین',
    'Lake Vermont HCC': 'زغال متالورژی مواد فرار کم لیک ورمانت',
    'Carborough Downs': 'زغال متالورژی مواد فرار کم کابروگ داونز',
    'Middlemount Coking': 'زغال متالورژی مواد فرار کم میدل مونت ککینگ',
    'Poitrel Semi Hard': 'زغال نیمه سخت مواد فرار کم پیترل',
    'Australia-China-Capesize': 'استرالیا-چین کیپ سایز',
    'Australia-Rotterdam-Capesize': 'استرالیا-نتردام کیپ سایز',
    'Australia-China-Panamax': 'استرالیا-چین پانامکس',
    'Australia-India-Panamax': 'استرالیا-هند پانامکس',
    'USEC-India-Panamax': 'ساحل شرقی آمریکا-هند پانامکس',
    'USEC-Rotterdam-Panamax': 'ساحل شرقی آمریکا-رتردام پانامکس',
    'USEC-Brazil-Panamax': 'ساحل شرقی آمریکا-برزیل پانامکس',
    'US Mobile-Rotterdam-Panamax': 'آمریکا-رتردام پانامکس'
}


excel_file_address_English = r'G:\Shared drives\Unlimited Drive\Global trading\Platts-Daily-Report\Platts-Data-English.xlsx'
export_to_excel(excel_file_address_English, dataframe_dict)
excel_format(excel_file_address_English)

for df_name in dataframe_dict:
    dataframe = dataframe_dict[df_name]
    translate_report(dataframe, translate_dict)

rule_columns_persian = ['تغییر', 'درصد تغییر']
currency_columns_persian = ['قیمت', 'تغییر', 'درصد تغییر']
percentage_column_persian = ['آهن', 'رطوبت',
                             'سیلیکا', 'آلومینا', 'فسفر', 'سولفور', 'شاخص سختی']

excel_file_address_Persian = r'G:\Shared drives\Unlimited Drive\Global trading\Platts-Daily-Report\Platts-Data-Persian.xlsx'
export_to_excel(excel_file_address_Persian, dataframe_dict)
excel_format(excel_file_address_Persian,percentage_list= percentage_column_persian, currency_list= currency_columns_persian,rule_columns=['تغییر','درصد تغییر'])